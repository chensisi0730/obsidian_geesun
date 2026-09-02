# -*- coding: utf-8 -*-
import os, sys, math, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from datetime import date
from pd1 import ROWS_P1, ROWS_P2
from pd2 import ROWS_P3, ROWS_P4
from pd3 import ROWS_P5, ROWS_P6, RISKS, INFO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUT = os.path.join(REPO_ROOT, "wiki/机器人/CATL 看机双足人形机器人开发计划.xlsx")
ROWS = ROWS_P1 + ROWS_P2 + ROWS_P3 + ROWS_P4 + ROWS_P5 + ROWS_P6
PROGRESS_OPTS = ["0%", "10%", "25%", "50%", "75%", "90%", "100%"]
STATUS_OPTS = ["未关闭", "缓解中", "已关闭"]

OWNER_MAP = {"A": "朱勇(A)", "B": "陈斯斯(B)", "C": "王朝(C)", "D": "杨海宾(D)", "E": "周子平(E)"}
OWNER_OPTS = list(OWNER_MAP.values()) + ["结构工程师（外部）"]

def fmt_owner(owner):
    """责任人代号 → 名字(代号)；多人组合保留原结构"""
    return re.sub(r"(?<![0-9A-Za-z])([ABCDE])(?![0-9A-Za-z])", lambda m: OWNER_MAP[m.group(1)], owner or "")

HDR_FILL = PatternFill("solid", fgColor="4472C4")
MIL_FILL = PatternFill("solid", fgColor="FFF2CC")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

def est_height(pairs):
    lines = 1
    for text, width in pairs:
        if not text:
            continue
        cap_units = max(8.0, float(width) - 2)
        n = 0
        for seg in str(text).split("\n"):
            units = sum(2 if ord(ch) > 127 else 1 for ch in seg)
            n += max(1, math.ceil(units / cap_units))
        lines = max(lines, n)
    return round(min(150.0, 16 + (lines - 1) * 13.5), 1)

def style_header(ws, row, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER_WRAP; cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w

def add_title(ws, ncols, text):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=text)
    t.font = TITLE_FONT
    ws.row_dimensions[1].height = 24

wb = openpyxl.Workbook()

# ---------- Sheet 1：开发计划 ----------
ws = wb.active
ws.title = "开发计划"
add_title(ws, 11, "CATL 看机双足人形机器人 — 开发计划（执行版，开始基线 2026-09-04）")
headers = ["序号", "关键阶段", "任务事项（汇总描述）", "开始日期", "结束日期", "责任人", "备注", "风险项目", "风险登记", "进度说明", "完成进度 %"]
widths  = [9, 26, 52, 11, 11, 20, 44, 38, 10, 16, 12]
HR = 3
style_header(ws, HR, headers, widths)

r = HR + 1
for row in ROWS:
    no, phase, task, ds, de, owner, note, risk, rid = row
    is_mil = no.endswith("-M")
    owner_disp = fmt_owner(owner)
    values = [no, phase, task, date.fromisoformat(ds), date.fromisoformat(de), owner_disp,
              note or None, risk or None, rid or None, "未开始", PROGRESS_OPTS[0]]
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        if c in (4, 5):
            cell.number_format = "yyyy-mm-dd"
            cell.alignment = Alignment(horizontal="center", vertical="top")
        elif c in (1, 9, 10, 11):
            cell.alignment = CENTER_WRAP
        else:
            cell.alignment = WRAP_TOP
        if is_mil:
            cell.fill = MIL_FILL
            if c == 3:
                cell.font = Font(bold=True)
    ws.row_dimensions[r].height = est_height([(task, widths[2]), (owner_disp, widths[5]), (note or "", widths[6]), (risk or "", widths[7])])
    r += 1

last = r - 1
ws.freeze_panes = f"A{HR+1}"
ws.auto_filter.ref = f"A{HR}:K{last}"
dv = DataValidation(type="list", formula1='"' + ",".join(PROGRESS_OPTS) + '"', allow_blank=True,
                    showInputMessage=True, promptTitle="完成进度 %",
                    prompt="从下拉列表选择：" + " / ".join(PROGRESS_OPTS))
ws.add_data_validation(dv)
dv.add(f"K{HR+1}:K{last}")
dvf = DataValidation(type="list", formula1='"' + ",".join(OWNER_OPTS) + '"', allow_blank=True,
                     showInputMessage=True, promptTitle="责任人",
                     prompt="从下拉列表选择（名字(代号)）：" + " / ".join(OWNER_OPTS))
ws.add_data_validation(dvf)
dvf.add(f"F{HR+1}:F{last}")

# ---------- Sheet 2：风险登记表 ----------
ws2 = wb.create_sheet("风险登记表")
add_title(ws2, 8, "CATL 看机项目 — 风险登记表（编号与规划工作簿风险管理表一致）")
h2 = ["风险ID", "风险项", "概率", "影响", "缓解措施", "关联任务（本表）", "责任阶段", "状态"]
w2 = [9, 30, 8, 8, 56, 40, 16, 12]
HR2 = 3
style_header(ws2, HR2, h2, w2)
r2 = HR2 + 1
for row in RISKS:
    for c, v in enumerate(row, 1):
        cell = ws2.cell(row=r2, column=c, value=v)
        cell.border = BORDER
        cell.alignment = WRAP_TOP if c in (5, 6) else Alignment(horizontal="center", vertical="top")
    ws2.row_dimensions[r2].height = est_height([(row[4], w2[4]), (row[5], w2[5])])
    r2 += 1
last2 = r2 - 1
ws2.freeze_panes = f"A{HR2+1}"
ws2.auto_filter.ref = f"A{HR2}:H{last2}"
dv2 = DataValidation(type="list", formula1='"' + ",".join(STATUS_OPTS) + '"', allow_blank=True,
                     showInputMessage=True, promptTitle="状态", prompt="从下拉列表选择：" + " / ".join(STATUS_OPTS))
ws2.add_data_validation(dv2)
dv2.add(f"H{HR2+1}:H{last2}")

# ---------- Sheet 3：说明与图例 ----------
ws3 = wb.create_sheet("说明与图例")
add_title(ws3, 2, "CATL 看机项目开发计划 — 说明与图例")
h3 = ["条目", "说明"]
w3 = [16, 105]
HR3 = 3
style_header(ws3, HR3, h3, w3)
r3 = HR3 + 1
for label, desc in INFO:
    c1 = ws3.cell(row=r3, column=1, value=label); c1.border = BORDER; c1.alignment = WRAP_TOP; c1.font = Font(bold=True)
    c2 = ws3.cell(row=r3, column=2, value=desc); c2.border = BORDER; c2.alignment = WRAP_TOP
    ws3.row_dimensions[r3].height = est_height([(desc, w3[1])])
    r3 += 1

wb.save(OUT)
print("saved:", OUT)

# ---------- 验证 ----------
wbv = openpyxl.load_workbook(OUT)
for s in wbv.worksheets:
    print(f"sheet={s.title} rows={s.max_row} cols={s.max_column}")
ws1 = wbv["开发计划"]
print("F4 =", ws1["F4"].value, "| F8 =", ws1["F8"].value, "| K4 =", ws1["K4"].value, "| D4 =", ws1["D4"].value, "| E4 =", ws1["E4"].value)
dvs = ws1.data_validations.dataValidation
print("devplan DV count:", len(dvs), "| range:", [str(d.sqref) for d in dvs], "| formula:", [d.formula1 for d in dvs])
ws2v = wbv["风险登记表"]
print("risk DV:", [(str(d.sqref), d.formula1) for d in ws2v.data_validations.dataValidation])
print("data rows:", last - HR, "| risk rows:", len(RISKS))
